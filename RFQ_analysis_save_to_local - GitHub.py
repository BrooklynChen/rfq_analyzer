#%%

import pandas as pd
import random
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import numpy as np
from sqlalchemy import create_engine
import urllib
import os
import joblib
import logging
from faker import Faker

# logging.basicConfig(filename='rfq_analysis.log', level=logging.DEBUG)
# logging.info("Starting script execution")

# # Download data from Access
# db_file = file_path
# username = username
# password = password

# # Set up the connection string with credentials
# connection_string = (
#     r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
#     f'DBQ={db_file};'
#     f'UID={username};'
#     f'PWD={password};'
# )

# # URL encode the connection string
# quoted_connection_string = urllib.parse.quote_plus(connection_string)

# # Create a SQLAlchemy engine using the connection string
# engine = create_engine(f'access+pyodbc:///?odbc_connect={quoted_connection_string}')

# # Execute a query to retrieve data from a specific table
# query_rfq = 'SELECT * FROM RFQ'
# query_rfqtofactory = 'SELECT * FROM RFQtoFactory'
# query_rfqquotesubmit = 'SELECT * FROM RFQQuoteSubmit'
# query_leadsinfo = 'SELECT * FROM "Leads Information"'

# try:
#     # Read the data directly into a pandas DataFrame
#     rfq = pd.read_sql(query_rfq, engine)
#     rfq_fac = pd.read_sql(query_rfqtofactory, engine)
#     rfq_q_submit = pd.read_sql(query_rfqquotesubmit, engine)
#     leads = pd.read_sql(query_leadsinfo, engine)


# except Exception as e:
#     print("Error:", e)
# finally:
#     # No need to explicitly close the engine; it handles connections
#     pass





fake = Faker()

# Generate mock data for the RFQ table
def generate_rfq_data(n=500):
    return pd.DataFrame([{
        'RFQID': f'RFQ{i}',
        'Customer': random.choice([f'Customer{i}' for i in range(1, 26)]),
        'RFQDate': fake.date_between(start_date='-5y', end_date='today'),
        'FactoryUsed': random.choice(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'A;B', 'E;G', 'B;D']),
        'AssignedSalesPerson': random.choice([
            'Randy Byerly', 'Robyn George', 'Alfred Ludwig',
            'James Aaron', 'Heath Mellen', 'Laura Jackson', 'David Tarver']),
        'Material': random.choice(['Fabric', 'Metal', 'Rubber', 'Plastic', 'Wire']),
        'ResultCode': random.choice(['Lost', 'Won', 'Pending', 'No Quote', np.nan]),
        'TotalAmount': round(random.uniform(1000, 100000), 2),
        'InquiryType': random.choice(['New Project', 'Price Update']),
        'Email': fake.email(),  # ✅ Fixed here
        'Quoted': random.choice([0,1])
    } for i in range(n)])

# Generate mock data for the RFQtoFactory table
def generate_rfq_to_factory(n=800):
    return pd.DataFrame([{
        'RFQID': f'RFQ{random.randint(0, 500)}',  # Foreign key to RFQ table
        'FactorySubmittedQuote': random.choice(['A', 'B', 'C', 'D', 'E', 'F', 'G', np.nan]),
        'FactoryQuoteRespone': random.choice(['Quoted', 'No Response', 'No Quote', 'Advised Drop', 'Estimate', 'Closed', np.nan])
    } for _ in range(n)])

# Generate mock data for the RFQQuoteSubmit table
def generate_rfq_quote_submit(n=700):
    return pd.DataFrame([{
        'RFQID': f'RFQ{random.randint(0, n)}',
        'PartNumber':f'PN{random.randint(1000, 9999)}',
        'Qty': random.randint(1, 10000),
        'Qty2': random.randint(1, 10000),
        'Qty3': random.randint(1, 10000),
        'Qty4': random.randint(1, 10000),
        'Qty5': random.randint(1, 10000),
        'PiecePrice': round(random.uniform(1.0, 10.0), 2),
        'PiecePrice2': round(random.uniform(1.0, 10.0), 2),
        'PiecePrice3': round(random.uniform(1.0, 10.0), 2),
        'PiecePrice4': round(random.uniform(1.0, 10.0), 2),
        'PiecePrice5': round(random.uniform(1.0, 10.0), 2),
        'PiecePrice6': round(random.uniform(1.0, 10.0), 2),
        'PiecePrice7': round(random.uniform(1.0, 10.0), 2),
        'PiecePrice8': round(random.uniform(1.0, 10.0), 2),
        'PiecePrice9': round(random.uniform(1.0, 10.0), 2),
        'PiecePrice10': round(random.uniform(1.0, 10.0), 2),
        'PiecePrice11': round(random.uniform(1.0, 10.0), 2),
        'PiecePrice12': round(random.uniform(1.0, 10.0), 2),
        'PiecePrice13': round(random.uniform(1.0, 10.0), 2),
        'PiecePrice14': round(random.uniform(1.0, 10.0), 2),
        'PiecePrice15': round(random.uniform(1.0, 10.0), 2),
    } for _ in range(n)])

# Generate all mock datasets
rfq = generate_rfq_data()
rfq['RFQ Due Date'] = rfq['RFQDate'] + timedelta(days=10)
rfq['FQUDate'] = rfq['RFQ Due Date'] + timedelta(days=3)
rfq_fac = generate_rfq_to_factory()
rfq_q_submit = generate_rfq_quote_submit()


# Clean RFQ file
rfq = rfq.rename(columns={'RFQDate':'Original RFQ Date', 'FQUDate':'Original Quote Date', 'AssignedSalesPerson':'Sales Rep'})
rfq = rfq.sort_values(by=['RFQID'])
rfq = rfq.replace({None: np.nan, '': np.nan})
rfq = rfq.apply(lambda x: np.nan if isinstance(x, str) and x.strip() == '' else x)
rfq['Customer'] = rfq['Customer'].str.upper()

# Add Year
rfq['Original RFQ Date'] = pd.to_datetime(rfq['Original RFQ Date'])
rfq['Original Quote Date'] = pd.to_datetime(rfq['Original Quote Date'])
rfq['Year'] = rfq['Original RFQ Date'].fillna(rfq['Original Quote Date'] - pd.Timedelta(days=5))
rfq['Year'] = rfq['Year'].fillna(rfq['RFQ Due Date'] - pd.Timedelta(days=5))
rfq['Year'] = rfq['Year'].dt.year
rfq['FactoryUsed'] = rfq['FactoryUsed'].fillna('Blanks')
rfq['ResultCode'] = rfq['ResultCode'].fillna('Blanks')

rfq = rfq.drop(columns=['Original Quote Date', 'Material', 'Email', 'Original RFQ Date', 'RFQ Due Date', 'TotalAmount'])

# Update Sales to the last rep for the customer
customer_to_sales = rfq.copy()
customer_to_sales_column = ['Customer', 'Sales Rep']
customer_to_sales = customer_to_sales[customer_to_sales_column]

customer_to_sales = customer_to_sales.groupby('Customer').last().reset_index()
customer_to_sales = customer_to_sales.fillna('Blanks Blanks')

# Merge new sales to the RFQ table
rfq = rfq.drop(columns=['Sales Rep'])
rfq = rfq.merge(customer_to_sales, on='Customer', how='left')

# Filter the dataset
rfq = rfq[rfq['InquiryType'] == 'New Project']
rfq = rfq[rfq['Year'] > 2019]
rfq = rfq[(rfq['Quoted'] == True)]
rfq['Sales Rep'] = rfq['Sales Rep'].fillna('Blanks Blanks')
rfq.to_excel('RFQ_clean.xlsx', index=False)

# Clean 'rfq_fac'
rfq_fac = rfq_fac.sort_values(by=['RFQID'])
rfq_fac = rfq_fac.fillna('Blanks')
rfq_fac = rfq_fac[rfq_fac['FactorySubmittedQuote'] != 'Blanks']
rfq_fac = rfq_fac.groupby(['RFQID', 'FactorySubmittedQuote']).last().reset_index()

# Merge 'rfq' and 'rfq_fac'
rfq2 = rfq.copy()
merged_df = rfq2.merge(rfq_fac, on ='RFQID', how='left')
merged_df_order =['RFQID', 'Customer', 'Year', 'Sales Rep', 'FactoryUsed', 'ResultCode', 'FactorySubmittedQuote', 'FactoryQuoteRespone']
merged_df = merged_df[merged_df_order]

# Split the rows where there's a comma and then explode the resulting lists into separate rows
merged_df['FactoryUsed'] = merged_df['FactoryUsed'].str.split(';')

# Explode the lists into individual rows (only works on columns that have lists or arrays)
rfq_clean_split = merged_df.explode('FactoryUsed', ignore_index=True)
rfq_clean_split = rfq_clean_split.fillna('Blanks')


# Calculate maximum WON amount of each RFQ
pd.options.display.float_format = '{:.3f}'.format

rfq_q_submit['Q1P1'] = rfq_q_submit['Qty'] * rfq_q_submit['PiecePrice']
rfq_q_submit['Q2P2'] = rfq_q_submit['Qty2'] * rfq_q_submit['PiecePrice2']
rfq_q_submit['Q3P3'] = rfq_q_submit['Qty3'] * rfq_q_submit['PiecePrice3']
rfq_q_submit['Q4P4'] = rfq_q_submit['Qty4'] * rfq_q_submit['PiecePrice4']
rfq_q_submit['Q5P5'] = rfq_q_submit['Qty5'] * rfq_q_submit['PiecePrice5']
rfq_q_submit['Q1P6'] = rfq_q_submit['Qty'] * rfq_q_submit['PiecePrice6']
rfq_q_submit['Q2P7'] = rfq_q_submit['Qty2'] * rfq_q_submit['PiecePrice7']
rfq_q_submit['Q3P8'] = rfq_q_submit['Qty3'] * rfq_q_submit['PiecePrice8']
rfq_q_submit['Q4P9'] = rfq_q_submit['Qty4'] * rfq_q_submit['PiecePrice9']
rfq_q_submit['Q5P10'] = rfq_q_submit['Qty5'] * rfq_q_submit['PiecePrice10']
rfq_q_submit['Q1P11'] = rfq_q_submit['Qty'] * rfq_q_submit['PiecePrice11']
rfq_q_submit['Q2P12'] = rfq_q_submit['Qty2'] * rfq_q_submit['PiecePrice12']
rfq_q_submit['Q3P13'] = rfq_q_submit['Qty3'] * rfq_q_submit['PiecePrice13']
rfq_q_submit['Q4P14'] = rfq_q_submit['Qty4'] * rfq_q_submit['PiecePrice14']
rfq_q_submit['Q5P15'] = rfq_q_submit['Qty5'] * rfq_q_submit['PiecePrice15']

rfq_q_submit.loc[:, 'Max_Amount'] = rfq_q_submit[['Q1P1', 'Q2P2', 'Q3P3', 'Q4P4', 'Q5P5', 
                                   'Q1P6', 'Q2P7', 'Q3P8', 'Q4P9', 
                                   'Q5P10', 'Q1P11', 'Q2P12', 'Q3P13', 
                                   'Q4P14', 'Q5P15']].max(axis=1)

rfq_q_submit['Max_Amount'] = rfq_q_submit['Max_Amount'].round(3)
rfq_q_submit = rfq_q_submit.groupby(['RFQID']).sum().reset_index()
rfq_q_submit_order = ['RFQID', 'Max_Amount']
rfq_q_submit = rfq_q_submit[rfq_q_submit_order]
################################################################

# Merge with RFQ to get [Customer. Year]
rfq_clean_2 = merged_df.copy()

rfq_won = rfq_clean_2[rfq_clean_2['ResultCode'] == 'Won']
rfq_won_order = ['RFQID', 'Customer', 'Year', 'FactoryUsed', 'Sales Rep']
rfq_won = rfq_won[rfq_won_order]
total_amount_won = rfq_won.merge(rfq_q_submit, on='RFQID', how='left')

def explode_factories(row):
    # Ensure 'FactoryUsed' is a string first
    if isinstance(row['FactoryUsed'], str):
        factories = row['FactoryUsed'].split(';')  # If it's a string, split it
    else:
        factories = row['FactoryUsed']  # If it's already a list, use it directly

    # Adjust the 'Max_Amount' accordingly if there are multiple factories
    num_factories = len(factories)
    if num_factories > 1:
        max_amounts = [row['Max_Amount'] / num_factories] * num_factories
    else:
        max_amounts = [row['Max_Amount']]

    # Create new rows with each factory and corresponding amount
    exploded_rows = pd.DataFrame({
        'RFQID': [row['RFQID']] * num_factories,
        'Customer': [row['Customer']] * num_factories,
        'Year': [row['Year']] * num_factories,
        'FactoryUsed': factories,
        'Sales Rep': [row['Sales Rep']] * num_factories,
        'Max_Amount': max_amounts
    })
    
    return exploded_rows


# Apply the function and concatenate the results
total_amount_won = pd.concat(total_amount_won.apply(explode_factories, axis=1).tolist(), ignore_index=True)
total_amount_won = total_amount_won.groupby(['RFQID', 'Customer', 'Year', 'FactoryUsed', 'Sales Rep']).last().reset_index()
total_amount_won.rename(columns={'FactoryUsed': 'Factory', 'Max_Amount': 'Total Amount'}, inplace=True)
total_amount_won.to_excel('Awarded_Amount_List.xlsx', index=False)

# FactoryUsed and Result
fac_used_result = rfq_clean_split.groupby(['RFQID','FactoryUsed']).last().reset_index()
fac_used_result = fac_used_result.groupby(['Year', 'FactoryUsed', 'ResultCode']).size().reset_index(name='Count')

# Pivot the dataframe so that 'ResultCode' values become columns
fac_used_result = fac_used_result.pivot_table(index=['FactoryUsed', 'Year'], 
                                columns='ResultCode', 
                                values='Count', 
                                aggfunc='sum', 
                                fill_value=0)

# Flatten the columns to a more readable format
fac_used_result.columns = [f'{col}' for col in fac_used_result.columns]
fac_used_result['Total Used'] = fac_used_result.sum(axis=1)

# Reset the index to make 'Factory' and 'Year' regular columns
fac_used_result = fac_used_result.reset_index()
fac_used_result = fac_used_result.rename(columns={'FactoryUsed':'Factory'})


# FactorySubmitted and Responses
fac_subm_response = rfq_clean_split.groupby(['RFQID','FactorySubmittedQuote']).last().reset_index()

# Group by 'FactorySubmittedQuote', 'Year', and 'FactoryQuoteRespone' and count the number of rows
fac_subm_response = fac_subm_response.groupby(['FactorySubmittedQuote', 'Year', 'FactoryQuoteRespone']).size().reset_index(name='Count')

# Pivot the dataframe so that 'FactoryQuoteRespone' values become columns
fac_subm_response = fac_subm_response.pivot_table(index=['FactorySubmittedQuote', 'Year'], 
                                columns='FactoryQuoteRespone', 
                                values='Count', 
                                aggfunc='sum', 
                                fill_value=0)

fac_subm_response.columns = [f'{col}' for col in fac_subm_response.columns]
fac_subm_response['Total Submitted Quote'] = fac_subm_response.sum(axis=1)

# Reset the index to make 'Factory' and 'Year' regular columns
fac_subm_response = fac_subm_response.reset_index()
fac_subm_response = fac_subm_response.rename(columns={'FactorySubmittedQuote':'Factory'})

# Merge
factory_ana = fac_subm_response.merge(fac_used_result, on=['Factory', 'Year'], how='outer')

new_order = ['Factory', 'Year', 'Total Submitted Quote', 'Quoted', 'No Quote_x', 'No Response', 'Advised Drop',
              'Estimate', 'Closed', 'Blanks_x', 'Total Used', 'Won', 'Lost', 'Pending', 'No Quote_y', 'Blanks_y']

factory_ana = factory_ana[new_order]
factory_ana = factory_ana.fillna(0)

current_date = datetime.now().strftime("%m.%d.%y")
factory_ana_name1 = f'RFQ_Analysis_by_Factory.xlsx'
factory_ana_name2 = f'RFQ_Analysis_by_Factory_{current_date}.xlsx'
factory_ana.to_excel(factory_ana_name1, index=False, float_format='%.3f')

# Format Excel file
wb = openpyxl.load_workbook(factory_ana_name1)
ws = wb.active

header_fill = PatternFill(start_color='204E84', end_color='204E84', fill_type='solid')
header_fill_2 = PatternFill(start_color='16365C', end_color='16365C', fill_type='solid')
header_font1 = Font(color='FFFFFF', bold=True)

max_row = ws.max_row
max_col = ws.max_column

# Fill color for header
for col in range(1, 11):
    header_cell = ws.cell(row=1, column=col)
    header_cell.fill = header_fill
    header_cell.font = header_font1
    header_cell.border = Border(left=None, right=None, top=None, bottom=None)

# Fill color for header
for col in range(11, max_col+1):
    header_cell = ws.cell(row=1, column=col)
    header_cell.fill = header_fill_2
    header_cell.font = header_font1
    header_cell.border = Border(left=None, right=None, top=None, bottom=None)

ws.insert_rows(1)
ws.merge_cells('A1:A2')
merged_cell = ws['A1']
merged_cell.value = 'Factory'
merged_cell.fill = header_fill
merged_cell.font = header_font1
merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws.merge_cells('B1:B2')
merged_cell = ws['B1']
merged_cell.value = 'Year'
merged_cell.fill = header_fill
merged_cell.font = header_font1
merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws.merge_cells('C1:J1')
merged_cell = ws['C1']
merged_cell.value = 'RFQ'
merged_cell.fill = header_fill
merged_cell.font = header_font1
merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws.merge_cells('K1:P1')
merged_cell = ws['K1']
merged_cell.value = 'Result'
merged_cell.fill = header_fill_2
merged_cell.font = header_font1
merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws.freeze_panes = 'A3'
ws.auto_filter.ref = f'A2:{chr(64 + max_col)}{ws.max_row}'
ws['A1'].fill = PatternFill(start_color='2A65AC', end_color='2A65AC', fill_type='solid')
ws['B1'].fill = PatternFill(start_color='2A65AC', end_color='2A65AC', fill_type='solid')

ws['C2'].font = Font(color='FFFF00', bold=True)
ws['K2'].font = Font(color='FFFF00', bold=True)
ws['E2'].value = 'No Quote'
ws['J2'].value = 'Blanks'
ws['O2'].value = 'No Quote'
ws['P2'].value = 'Blanks'

# Change column name
for cell in ws[1]:  # Loop through the first row (header)
    if cell.value == 'Total Amount':
        cell.value = 'Total Amount (Reference)'
        break

wb.save(factory_ana_name2)

bi_fac = pd.read_excel(factory_ana_name1)
bi_fac_order = ['Factory', 'Year', 'Total Submitted Quote', 'Total Used', 'Won']
bi_fac = bi_fac[bi_fac_order]
bi_fac['Win Rate'] = bi_fac['Won']/bi_fac['Total Used']

# Calculate Usage Rate
bi_fac['Usage Rate'] = bi_fac['Total Used'] / bi_fac['Total Submitted Quote']

total_amount_won_2 = total_amount_won.copy()
total_amount_won_2 = total_amount_won_2.drop(columns=['Customer', 'Sales Rep', 'RFQID'])
total_amount_won_2 = total_amount_won_2.groupby(['Year', 'Factory']).sum().reset_index()

bi_fac = bi_fac.merge(total_amount_won_2, on=['Year', 'Factory'], how='left')
bi_fac = bi_fac.fillna(0)
bi_fac = bi_fac[bi_fac['Factory'] != 'Blanks']
bi_fac.to_excel(factory_ana_name1, index=False, float_format='%.3f')


# RFQ Analysis by Customer
cus_ana = rfq_clean_split.copy()
cus_subm_result = cus_ana.groupby(['RFQID', 'Customer', 'Year']).last().reset_index()
cus_subm_result = cus_subm_result.groupby(['Customer', 'Year', 'ResultCode']).size().reset_index(name='Count')
cus_subm_result = cus_subm_result.pivot_table(index=['Customer', 'Year'], 
                            columns=['ResultCode'], 
                            values='Count', 
                            aggfunc='sum', 
                            fill_value=0)

cus_subm_result = cus_subm_result.reset_index()
cus_subm_result.rename(columns={'FactoryUsed': 'Factory'}, inplace=True)
cus_subm_result = cus_subm_result.merge(customer_to_sales, on='Customer', how='left')

cus_amount = total_amount_won.copy()
cus_amount = cus_amount.drop(columns=['RFQID', 'Factory', 'Sales Rep'])
cus_amount = cus_amount.groupby(['Customer', 'Year']).sum().reset_index()
cus_subm_result = cus_subm_result.merge(cus_amount, on=['Customer', 'Year'], how='left')

cus_subm_result_1 = cus_subm_result.copy()
cus_subm_result_order = ['Customer', 'Year', 'Sales Rep', 'Won', 'Lost', 'No Quote', 'Pending', 'Blanks']
cus_subm_result_1 = cus_subm_result_1[cus_subm_result_order]

cus_total2 = rfq_clean_split.copy()
cus_total2 = cus_total2.groupby(['RFQID']).last().reset_index()
cus_total2 = cus_total2.groupby(['Customer', 'Year']).size().reset_index(name='Total Submitted RFQs')

cus_total3 = rfq_clean_split.copy()
cus_total3 = cus_total3.groupby(['RFQID']).last().reset_index()
cus_total3 = cus_total3[cus_total3['ResultCode'] != 'No Quote']
cus_total3 = cus_total3.groupby(['Customer', 'Year']).size().reset_index(name='Total Quoted RFQs')

cus_subm_result_1 = cus_subm_result_1.merge(cus_total2, on=['Customer', 'Year'], how='outer')
cus_subm_result_1 = cus_subm_result_1.merge(cus_total3, on=['Customer', 'Year'], how='outer')
cus_subm_result_1 = cus_subm_result_1.fillna(0)

total_amount_won_4 = total_amount_won.copy()
total_amount_won_4 = total_amount_won_4.drop(columns=['Sales Rep', 'Factory', 'RFQID'])
total_amount_won_4 = total_amount_won_4.groupby(['Year', 'Customer']).sum().reset_index()

cus_subm_result_1 = cus_subm_result_1.merge(total_amount_won_4, on=['Year', 'Customer'], how='left')
cus_subm_result_1['Total Amount'] = cus_subm_result_1['Total Amount'].fillna(0)
cus_subm_result_1['Win Rate'] = cus_subm_result_1['Won']/cus_subm_result_1['Total Quoted RFQs']
cus_subm_result_1 = cus_subm_result_1.fillna(0)

cus_subm_result_order = ['Customer', 'Year', 'Sales Rep', 'Total Submitted RFQs', 'Total Quoted RFQs', 'Total Amount', 'Win Rate', 'Won', 'Lost', 'No Quote', 'Pending', 'Blanks']
cus_subm_result_1 = cus_subm_result_1[cus_subm_result_order]
pivot_cus_name1 = f'RFQ_Analysis_by_Customer.xlsx'
pivot_cus_name2 = f'RFQ_Analysis_by_Customer_{current_date}.xlsx'

cus_subm_result_1.to_excel(pivot_cus_name1, index=False, float_format='%.3f')
cus_number_by_sales_by_year = cus_subm_result_1.copy()

cus_list_ = ['Customer']
customer_list = cus_subm_result_1[cus_list_]
customer_list = customer_list.drop_duplicates(subset='Customer', keep='last')
customer_list.to_excel('Customer_List.xlsx', index=False)

# Format Excel file
wb = openpyxl.load_workbook(pivot_cus_name1)
ws = wb.active

header_fill_1 = PatternFill(start_color='19896E', end_color='19896E', fill_type='solid')
header_fill_3 = PatternFill(start_color='136753', end_color='136753', fill_type='solid')
header_font1 = Font(color='FFFFFF', bold=True)

max_row = ws.max_row
max_col = ws.max_column

# Fill color for header
for col in range(1, max_col+1):
    header_cell = ws.cell(row=1, column=col)
    header_cell.fill = header_fill_3
    header_cell.font = header_font1
    header_cell.border = Border(left=None, right=Side(style='thin', color='D9D9D9'), top=None, bottom=None)

ws.insert_rows(1)
ws.merge_cells('A1:A2')
merged_cell = ws['A1']
merged_cell.value = 'Customer'
merged_cell.fill = header_fill_1
merged_cell.font = header_font1
merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws.merge_cells('B1:B2')
merged_cell = ws['B1']
merged_cell.value = 'Year'
merged_cell.fill = header_fill_1
merged_cell.font = header_font1
merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws.merge_cells('C1:C2')
merged_cell = ws['C1']
merged_cell.value = 'Sales Rep'
merged_cell.fill = header_fill_1
merged_cell.font = header_font1
merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws.merge_cells('D1:D2')
merged_cell = ws['D1']
merged_cell.value = 'Total Submitted RFQs'
merged_cell.fill = header_fill_1
merged_cell.font = header_font1
merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws.merge_cells('E1:E2')
merged_cell = ws['E1']
merged_cell.value = 'Total Quoted RFQs'
merged_cell.fill = header_fill_1
merged_cell.font = header_font1
merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws.merge_cells('F1:F2')
merged_cell = ws['F1']
merged_cell.value = 'Total Amount'
merged_cell.fill = header_fill_1
merged_cell.font = header_font1
merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws.merge_cells('G1:G2')
merged_cell = ws['G1']
merged_cell.value = 'Win Rate'
merged_cell.fill = header_fill_1
merged_cell.font = header_font1
merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)



ws.merge_cells('H1:L1')
merged_cell = ws['H1']
merged_cell.value = 'Result'
merged_cell.fill = header_fill_3
merged_cell.font = header_font1
merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws.freeze_panes = 'A3'
ws.auto_filter.ref = f'A2:{chr(64 + max_col)}{ws.max_row}'

# Change column name
for cell in ws[1]:  # Loop through the first row (header)
    if cell.value == 'Total Amount':
        cell.value = 'Total Amount (Reference)'
        break
    
wb.save(pivot_cus_name2)


# Customer FactoryUsed and Results + FacSubm and Response
cus_submfac_response = rfq_clean_split.groupby(['RFQID', 'Customer', 'Year', 'FactorySubmittedQuote', 'FactoryQuoteRespone']).last().reset_index()
cus_submfac_response = cus_submfac_response.groupby(['Customer', 'Year', 'FactorySubmittedQuote', 'FactoryQuoteRespone']).size().reset_index(name='Count')

# Pivot the DataFrame, with 'FactoryUsed' as columns and 'ResultCode' as another dimension in the columns
cus_submfac_response = cus_submfac_response.pivot_table(index=['Customer', 'Year', 'FactorySubmittedQuote'], 
                            columns=['FactoryQuoteRespone'], 
                            values='Count', 
                            aggfunc='sum', 
                            fill_value=0)

cus_submfac_response = cus_submfac_response.reset_index()
cus_submfac_response['Total Submitted RFQs'] = cus_submfac_response.loc[:, 'Advised Drop':'Quoted'].sum(axis=1)


cus_usedfac_result = rfq_clean_split.groupby(['RFQID', 'Customer', 'Year', 'FactoryUsed', 'ResultCode']).last().reset_index()
cus_usedfac_result = cus_usedfac_result.groupby(['Customer', 'Year', 'FactoryUsed', 'ResultCode']).size().reset_index(name='Count')

# Pivot the DataFrame, with 'FactoryUsed' as columns and 'ResultCode' as another dimension in the columns
cus_usedfac_result = cus_usedfac_result.pivot_table(index=['Customer', 'Year', 'FactoryUsed'], 
                            columns=['ResultCode'], 
                            values='Count', 
                            aggfunc='sum', 
                            fill_value=0)

cus_usedfac_result = cus_usedfac_result.reset_index()

cus_usedfac_result['Total Quoted'] = cus_usedfac_result['Blanks']+cus_usedfac_result['Lost']+cus_usedfac_result['Pending']+cus_usedfac_result['Won']
cus_usedfac_result = cus_usedfac_result.rename(columns={'FactoryUsed': 'Factory'})

total_amount_won_5 = total_amount_won.copy()
total_amount_won_5 = total_amount_won_5.drop(columns=['RFQID', 'Sales Rep'])
total_amount_won_5 = total_amount_won_5.groupby(['Year', 'Customer', 'Factory']).sum().reset_index()

cus_usedfac_result = cus_usedfac_result.merge(total_amount_won_5, on=['Year', 'Customer', 'Factory'], how='left')
cus_usedfac_result['Total Amount'] = cus_usedfac_result['Total Amount'].fillna(0)

cus_usedfac_result = cus_usedfac_result.merge(customer_to_sales, on='Customer', how='left')
cus_usedfac_result['Win Rate'] = cus_usedfac_result['Won']/cus_usedfac_result['Total Quoted']

cus_submfac_response_ord = ['Customer', 'Year', 'FactorySubmittedQuote', 'Total Submitted RFQs']
cus_submfac_response = cus_submfac_response[cus_submfac_response_ord]
cus_submfac_response = cus_submfac_response.rename(columns={'FactorySubmittedQuote': 'Factory'})
cus_usedfac_result = cus_usedfac_result.merge(cus_submfac_response, on=['Year', 'Customer', 'Factory'], how='outer')
cus_usedfac_result = cus_usedfac_result.drop(columns=['Sales Rep'])
cus_usedfac_result = cus_usedfac_result.merge(customer_to_sales, on='Customer', how='left')

reset_order = ['Customer', 'Year', 'Factory', 'Sales Rep',
       'Total Submitted RFQs', 'Total Quoted', 'Total Amount',
       'Won', 'Blanks', 'Lost', 'No Quote', 'Pending', 'Win Rate']
cus_usedfac_result = cus_usedfac_result[reset_order]
cus_usedfac_result = cus_usedfac_result.fillna(0)

cus_usedfac_result.to_excel('Customer_UsedFactory_Result.xlsx', index=False)


# Sales to Factories
sales_submit = rfq_clean_split.copy()
sales_submit = sales_submit.groupby(['RFQID', 'FactorySubmittedQuote']).last().reset_index()
sales_submit = sales_submit.groupby(['Sales Rep', 'Year', 'FactorySubmittedQuote','FactoryQuoteRespone']).size().reset_index(name='Count')

# Pivot the DataFrame, with 'FactoryUsed' as columns and 'ResultCode' as another dimension in the columns
sales_submit = sales_submit.pivot_table(index=['Sales Rep', 'Year', 'FactorySubmittedQuote'], 
                            columns=['FactoryQuoteRespone'], 
                            values='Count', 
                            aggfunc='sum', 
                            fill_value=0)
sales_submit = sales_submit.reset_index()

sales_received = rfq_clean_split.copy()
sales_received = sales_received.groupby(['RFQID', 'FactoryUsed']).last().reset_index()

sales_received = sales_received.groupby(['Sales Rep', 'Year', 'FactoryUsed', 'ResultCode']).size().reset_index(name='Count')

# Pivot the DataFrame, with 'FactoryUsed' as columns and 'ResultCode' as another dimension in the columns
sales_received = sales_received.pivot_table(index=['Sales Rep', 'Year', 'FactoryUsed'], 
                            columns=['ResultCode'], 
                            values='Count', 
                            aggfunc='sum', 
                            fill_value=0)
sales_received = sales_received.reset_index()


# Sales Rep Analysis
sales_result = rfq_clean_split.copy()
sales_result = sales_result.groupby(['RFQID']).last().reset_index()
sales_result = sales_result.groupby(['Sales Rep', 'Year', 'ResultCode']).size().reset_index(name='Count')

# Pivot the dataframe so that 'Customer Result' values become columns
sales_result = sales_result.pivot_table(index=['Sales Rep', 'Year'], 
                                columns='ResultCode', 
                                values='Count', 
                                aggfunc='sum', 
                                fill_value=0)
sales_result = sales_result.reset_index()

sales_received2 = rfq_clean_split.copy()
sales_received2 = sales_received2.groupby(['RFQID']).last().reset_index()
sales_received2 = sales_received2.groupby(['Sales Rep', 'Year']).size().reset_index(name='Total Received RFQs')

sales_received3 = rfq_clean_split.copy()
sales_received3 = sales_received3.groupby(['RFQID']).last().reset_index()
sales_received3 = sales_received3[sales_received3['ResultCode'] != 'No Quote']
sales_received3 = sales_received3.groupby(['Sales Rep', 'Year']).size().reset_index(name='Total Quoted RFQs')

sales_ana = sales_received2.merge(sales_received3, on=['Sales Rep', 'Year'], how='outer')
sales_ana = sales_ana.merge(sales_result, on=['Sales Rep', 'Year'], how='outer')

sales_ana['Win Rate'] = sales_ana['Won']/sales_ana['Total Quoted RFQs']

total_amount_won_3 = total_amount_won.copy()
total_amount_won_3 = total_amount_won_3.drop(columns=['Customer', 'Factory', 'RFQID'])
total_amount_won_3 = total_amount_won_3.groupby(['Year', 'Sales Rep']).sum().reset_index()

sales_ana = sales_ana.merge(total_amount_won_3, on=['Year', 'Sales Rep'], how='left')
sales_ana['Total Amount'] = sales_ana['Total Amount'].fillna(0)

# Calculate the number of customer of the sales by year
keep_columns = ['Customer', 'Year', 'Sales Rep']
cus_number_by_sales_by_year = cus_number_by_sales_by_year[keep_columns]

new_cus = cus_number_by_sales_by_year.copy()
new_cus['Year'] = new_cus['Year'].astype(int)

cus_number_by_sales_by_year = cus_number_by_sales_by_year.groupby(['Sales Rep', 'Year']).size().reset_index(name='Number of Customers')
sales_ana = sales_ana.merge(cus_number_by_sales_by_year, on=['Sales Rep', 'Year'], how='left')


# Define a function to calculate new customers for each Sales Rep
def calculate_new_customers(df):
    # Initialize an empty list to store the results
    results = []

    # Loop through the years 2021 to 2025
    for year in range(2025, max(df['Year']) + 1):
        # Filter the data for the current year
        current_year_data = df[df['Year'] == year]

        # Filter the data for the previous year
        prev_year_data = df[df['Year'] == (year - 1)]

        # Get the unique customers for the current and previous years
        current_year_customers = set(current_year_data['Customer'])
        prev_year_customers = set(prev_year_data['Customer'])

        # Find customers in the current year but not in the previous year
        new_customers = current_year_customers - prev_year_customers

        # For each Sales Rep, count the number of new customers
        for rep in current_year_data['Sales Rep'].unique():
            rep_customers = current_year_data[current_year_data['Sales Rep'] == rep]['Customer']
            # Calculate how many of the customers in this rep's list are new
            new_customers_for_rep = len(set(rep_customers) & new_customers)
            results.append({'Year': year, 'Sales Rep': rep, 'New Customers': new_customers_for_rep})
    
    # Return the results as a DataFrame
    return pd.DataFrame(results)

# Call the function to calculate new customers
new_customers = calculate_new_customers(new_cus)

sales_ana = sales_ana.merge(new_customers, on=['Sales Rep', 'Year'], how='left')
sales_ana = sales_ana.fillna(0)
sales_ana.to_excel(f'RFQ_Analysis_by_Sales.xlsx', index=False, float_format='%.3f')


# Sales List
sales_list_order = ['Sales Rep']
sales_list = customer_to_sales[sales_list_order]
sales_list = sales_list.drop_duplicates(subset='Sales Rep', keep='last')
sales_list = sales_list.fillna('N A')

# Define the abbreviation function
def abbrevName(name):
    # Split the name by spaces to get the first and last names
    names = name.split()
    # Return the initials of the first and last names in uppercase
    return f"{names[0][0]}{names[1][0]}".upper()

# Apply the function to the 'Sales Rep' column and create a new 'Sales Rep Initials' column
sales_list['Sales Code'] = sales_list['Sales Rep'].apply(abbrevName)
sales_list =sales_list.rename(columns={'Sales Rep': 'Sales Name'})

sales_list.to_excel(f'Sales_List.xlsx', index=False)



# Win Rate Predictions
df = cus_usedfac_result.copy()
df['Total Submitted RFQs'] = pd.to_numeric(df['Total Submitted RFQs'], errors='coerce')
df['Total Quoted'] = pd.to_numeric(df['Total Quoted'], errors='coerce')
df['Blanks'] = pd.to_numeric(df['Blanks'], errors='coerce')
df['Lost'] = pd.to_numeric(df['Lost'], errors='coerce')
df['No Quote'] = pd.to_numeric(df['No Quote'], errors='coerce')
df['Pending'] = pd.to_numeric(df['Pending'], errors='coerce')
df['Won'] = pd.to_numeric(df['Won'], errors='coerce')
df['Total Amount'] = pd.to_numeric(df['Total Amount'], errors='coerce')

df = df.drop(columns=['Win Rate', 'Year', 'Sales Rep'])
df = df.groupby(['Customer', 'Factory']).sum().reset_index()
df['Win Rate'] = df['Won']/df['Total Quoted']
df['Customer-Factory'] = df['Customer'].astype(str) + '-' + df['Factory'].astype(str)

# Create customer-level and factory-level feature
df2 = df.copy()

customer_avg = df2.groupby('Customer').sum(numeric_only=True).reset_index()
customer_avg['Cus_Win Rate'] = customer_avg['Won']/customer_avg['Total Quoted']
customer_avg = customer_avg.drop(columns=['Win Rate', 'Blanks', 'Lost', 'No Quote', 'Pending'])
customer_avg = customer_avg.rename(columns={'Total Submitted RFQs': 'Cus_Total Submitted RFQs', 'Total Quoted':'Cus_Total Quoted', 'Total Amount':'Cus_Total Amount', 'Won':'Cus_Won'})

factory_avg = df2.groupby('Factory').sum(numeric_only=True).reset_index()
factory_avg['Fac_Win Rate'] = factory_avg['Won']/factory_avg['Total Quoted']
factory_avg = factory_avg.drop(columns=['Win Rate', 'Blanks', 'Lost', 'No Quote', 'Pending'])
factory_avg = factory_avg.rename(columns={'Total Submitted RFQs': 'Fac_Total Submitted RFQs', 'Total Quoted':'Fac_Total Quoted', 'Total Amount':'Fac_Total Amount', 'Won':'Fac_Won'})


# Merge these averages back into the original data
df = df.merge(customer_avg, on='Customer')
df = df.merge(factory_avg, on='Factory')

df['Cus_Win Rate1'] = np.log1p(df['Cus_Win Rate'])
df = df.fillna(0)

rf_best = joblib.load(r'C:\data\Project12_RFQ_Analysis\best_model.pkl')

X = df[['Won', 'Cus_Win Rate1', 'Cus_Won', 'Fac_Won', 'Fac_Win Rate', 'Total Submitted RFQs']]

log_pred_all = rf_best.predict(X)

y_pred_all = np.expm1(log_pred_all)

df_new = df.copy()
df_new['Actual Win Rate'] = df['Win Rate']
df_new['Predicted Win Rate'] = y_pred_all
df_new.to_excel('Win_Rate_Predictions.xlsx', index=False)


# %%
