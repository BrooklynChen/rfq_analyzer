import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Alignment

# Format Excel file
def format_factory_excel(factory_ana_name1, factory_ana_name2, rfq_won):
    wb = openpyxl.load_workbook(factory_ana_name1)
    ws = wb.active

    header_fill = PatternFill(start_color='204E84', end_color='204E84', fill_type='solid')
    header_fill_2 = PatternFill(start_color='16365C', end_color='16365C', fill_type='solid')
    header_font1 = Font(color='FFFFFF', bold=True)

    max_row = ws.max_row
    max_col = ws.max_column

    # Fill color for header
    for col in range(1, 11):
        header_cell = ws.cell(row=1, column=col)
        header_cell.fill = header_fill
        header_cell.font = header_font1
        header_cell.border = Border(left=None, right=None, top=None, bottom=None)

    # Fill color for header
    for col in range(11, max_col+1):
        header_cell = ws.cell(row=1, column=col)
        header_cell.fill = header_fill_2
        header_cell.font = header_font1
        header_cell.border = Border(left=None, right=None, top=None, bottom=None)

    ws.insert_rows(1)
    ws.merge_cells('A1:A2')
    merged_cell = ws['A1']
    merged_cell.value = 'Factory'
    merged_cell.fill = header_fill
    merged_cell.font = header_font1
    merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.merge_cells('B1:B2')
    merged_cell = ws['B1']
    merged_cell.value = 'Year'
    merged_cell.fill = header_fill
    merged_cell.font = header_font1
    merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.merge_cells('C1:J1')
    merged_cell = ws['C1']
    merged_cell.value = 'RFQ'
    merged_cell.fill = header_fill
    merged_cell.font = header_font1
    merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.merge_cells('K1:P1')
    merged_cell = ws['K1']
    merged_cell.value = 'Result'
    merged_cell.fill = header_fill_2
    merged_cell.font = header_font1
    merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{chr(64 + max_col)}{ws.max_row}'
    ws['A1'].fill = PatternFill(start_color='2A65AC', end_color='2A65AC', fill_type='solid')
    ws['B1'].fill = PatternFill(start_color='2A65AC', end_color='2A65AC', fill_type='solid')

    ws['C2'].font = Font(color='FFFF00', bold=True)
    ws['K2'].font = Font(color='FFFF00', bold=True)
    ws['E2'].value = 'No Quote'
    ws['J2'].value = 'Blanks'
    ws['O2'].value = 'No Quote'
    ws['P2'].value = 'Blanks'

    # Rename column
    for cell in ws[1]:  # Loop through the first row (header)
        if cell.value == 'Total Amount':
            cell.value = 'Total Amount (Reference)'
            break

    wb.save(factory_ana_name2)

    bi_fac = pd.read_excel(factory_ana_name1)
    bi_fac_order = ['Factory', 'Year', 'Total Submitted Quote', 'Total Used', 'Won']
    bi_fac = bi_fac[bi_fac_order]
    bi_fac['Win Rate'] = bi_fac['Won']/bi_fac['Total Used']

    # Calculate Usage Rate
    bi_fac['Usage Rate'] = bi_fac['Total Used'] / bi_fac['Total Submitted Quote']

    total_amount_won_2 = rfq_won.copy()
    total_amount_won_2 = total_amount_won_2.drop(columns=['Customer', 'Sales Rep', 'RFQID'])
    total_amount_won_2 = total_amount_won_2.groupby(['Year', 'Factory']).sum().reset_index()

    bi_fac = bi_fac.merge(total_amount_won_2, on=['Year', 'Factory'], how='left')
    bi_fac = bi_fac.fillna(0)
    bi_fac = bi_fac[bi_fac['Factory'] != 'Blanks']
    bi_fac.to_excel(factory_ana_name1, index=False, float_format='%.3f')