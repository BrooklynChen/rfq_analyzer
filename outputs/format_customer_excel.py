import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Alignment, Side

def format_customer_excel(pivot_cus_name1, pivot_cus_name2):
    wb = openpyxl.load_workbook(pivot_cus_name1)
    ws = wb.active

    header_fill_1 = PatternFill(start_color='19896E', end_color='19896E', fill_type='solid')
    header_fill_3 = PatternFill(start_color='136753', end_color='136753', fill_type='solid')
    header_font1 = Font(color='FFFFFF', bold=True)

    max_row = ws.max_row
    max_col = ws.max_column

    # Fill color for header
    for col in range(1, max_col+1):
        header_cell = ws.cell(row=1, column=col)
        header_cell.fill = header_fill_3
        header_cell.font = header_font1
        header_cell.border = Border(left=None, right=Side(style='thin', color='D9D9D9'), top=None, bottom=None)

    ws.insert_rows(1)
    ws.merge_cells('A1:A2')
    merged_cell = ws['A1']
    merged_cell.value = 'Customer'
    merged_cell.fill = header_fill_1
    merged_cell.font = header_font1
    merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.merge_cells('B1:B2')
    merged_cell = ws['B1']
    merged_cell.value = 'Year'
    merged_cell.fill = header_fill_1
    merged_cell.font = header_font1
    merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.merge_cells('C1:C2')
    merged_cell = ws['C1']
    merged_cell.value = 'Sales Rep'
    merged_cell.fill = header_fill_1
    merged_cell.font = header_font1
    merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.merge_cells('D1:D2')
    merged_cell = ws['D1']
    merged_cell.value = 'Total Submitted RFQs'
    merged_cell.fill = header_fill_1
    merged_cell.font = header_font1
    merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.merge_cells('E1:E2')
    merged_cell = ws['E1']
    merged_cell.value = 'Total Quoted RFQs'
    merged_cell.fill = header_fill_1
    merged_cell.font = header_font1
    merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.merge_cells('F1:F2')
    merged_cell = ws['F1']
    merged_cell.value = 'Total Amount'
    merged_cell.fill = header_fill_1
    merged_cell.font = header_font1
    merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.merge_cells('G1:G2')
    merged_cell = ws['G1']
    merged_cell.value = 'Win Rate'
    merged_cell.fill = header_fill_1
    merged_cell.font = header_font1
    merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)



    ws.merge_cells('H1:L1')
    merged_cell = ws['H1']
    merged_cell.value = 'Result'
    merged_cell.fill = header_fill_3
    merged_cell.font = header_font1
    merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{chr(64 + max_col)}{ws.max_row}'

    # Change column name
    for cell in ws[1]:  # Loop through the first row (header)
        if cell.value == 'Total Amount':
            cell.value = 'Total Amount (Reference)'
            break
        
    wb.save(pivot_cus_name2)